"""
Cognitive & Emotional Performance Research Dashboard
Research-Grade Analysis System for Cognitive Load & Performance Studies

Cognitive Metrics:
- Understanding Index (U_C, U_W, U_NA)
- Debugging Index (D_C, D_W, D_NA)
- Completion Index (C_C, C_W, C_NA)
- Technical Comprehensive Index
- Emotion Index (Emotion_Value_Scaled)
- Cognitive Load (CL1, CL2)
- Problem Performance (P1-P7 Before/After)

Run: python app.py
"""

from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from datetime import datetime
import os
import pandas as pd
import numpy as np
import secrets
import json
import io
import re

# ==================== FLASK APP INITIALIZATION ====================

app = Flask(__name__)
app.config['SECRET_KEY'] = secrets.token_hex(32)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///cogn_system.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024
app.config['UPLOAD_FOLDER'] = os.path.abspath(os.path.join(os.path.dirname(__file__), 'uploads'))

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
db = SQLAlchemy(app)

# ==================== DATABASE MODELS ====================

class User(db.Model):
    """Regular user model"""
    __tablename__ = 'users'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    is_active = db.Column(db.Boolean, default=True)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Admin(db.Model):
    """Administrator model"""
    __tablename__ = 'admins'
    
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    full_name = db.Column(db.String(100), nullable=False)
    role = db.Column(db.String(50), default='admin')
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    last_login = db.Column(db.DateTime)
    is_active = db.Column(db.Boolean, default=True)
    
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class DataFile(db.Model):
    """Uploaded data files tracking"""
    __tablename__ = 'data_files'
    
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(200), unique=True, nullable=False)
    original_name = db.Column(db.String(200), nullable=False)
    file_path = db.Column(db.String(300), nullable=False)
    uploaded_by = db.Column(db.String(80), nullable=False)
    file_size = db.Column(db.Integer)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    is_active = db.Column(db.Boolean, default=True)
    description = db.Column(db.Text)
    rows_count = db.Column(db.Integer)
    columns_count = db.Column(db.Integer)


class Analysis(db.Model):
    """Published analysis results"""
    __tablename__ = 'analyses'
    
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    data_file_id = db.Column(db.Integer, db.ForeignKey('data_files.id'), nullable=False)
    created_by = db.Column(db.String(80), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Relationships
    data_file = db.relationship('DataFile', backref='analyses')
    shared_users = db.relationship('User', secondary='user_analysis_permission', backref='shared_analyses')
    
    # Analysis metrics
    metrics_json = db.Column(db.Text)  # JSON with all calculations
    visualizations_json = db.Column(db.Text)  # JSON with chart data
    
    is_published = db.Column(db.Boolean, default=False)
    published_at = db.Column(db.DateTime)


class UserAnalysisPermission(db.Model):
    """Many-to-many relationship for sharing analyses with users"""
    __tablename__ = 'user_analysis_permission'
    
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    analysis_id = db.Column(db.Integer, db.ForeignKey('analyses.id'), nullable=False)
    shared_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    __table_args__ = (db.UniqueConstraint('user_id', 'analysis_id', name='unique_user_analysis'),)


class TechnicalAnalysis(db.Model):
    """Technical analysis with Bloom's taxonomy weighting"""
    __tablename__ = 'technical_analyses'
    
    id = db.Column(db.Integer, primary_key=True)
    title = db.Column(db.String(200), nullable=False)
    description = db.Column(db.Text)
    uploaded_file_path = db.Column(db.String(300), nullable=False)
    uploaded_by = db.Column(db.String(80), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    
    # Source data info
    original_filename = db.Column(db.String(200), nullable=False)
    rows_count = db.Column(db.Integer)
    
    # Enriched dataset (computed)
    enriched_file_path = db.Column(db.String(300))  # Path to Excel with T1, T2, T3, CL1 columns
    
    # Visualization data
    viz_data_json = db.Column(db.Text)  # Chart data (bar, stacked, radar, boxplot)
    
    # Validation status
    validation_status = db.Column(db.String(50), default='pending')  # pending, valid, invalid
    validation_errors = db.Column(db.Text)  # JSON list of errors if invalid


# ==================== TECHNICAL ANALYSIS VALIDATION & PROCESSING ====================

class PaperCompliantAnalyzer:
    """
    Paper-compliant analyzer implementing:
    "An empirical approach to investigate the impact of technical and 
    non-technical parameters on programmers' code comprehension proficiency"
    
    Expects 6 pre-computed indices from Excel (all in [0,1]):
    - Technical: CUI, CDI, CCI
    - Non-Technical: PSI, EI, LTMI
    
    Computes:
    - TI = Average((CUI × 0.6 scaled) + (CDI × 0.2 scaled) + (CCI × 0.2 scaled))
    - NTI = (PSI + EI + LTMI) / 3
    - CMI_P = 0.5 × TI + 0.5 × NTI
    """
    
    def __init__(self, df):
        self.df = df
        self.required_cols = ['CUI', 'CDI', 'CCI', 'PSI', 'EI', 'LTMI']
        self.errors = []
        self.warnings = []
    
    def validate(self):
        """Validate all 6 required indices are present and in [0,1] range"""
        missing = [col for col in self.required_cols if col not in self.df.columns]
        if missing:
            self.errors.append(f"Missing required indices: {', '.join(missing)}")
            return False
        
        # Validate each column is in [0, 1] range
        for col in self.required_cols:
            try:
                if (self.df[col] < 0).any() or (self.df[col] > 1).any():
                    self.errors.append(
                        f"Column '{col}' has values outside [0,1] range "
                        f"(min={self.df[col].min():.4f}, max={self.df[col].max():.4f})"
                    )
                
                if self.df[col].isnull().any():
                    self.errors.append(
                        f"Column '{col}' contains {self.df[col].isnull().sum()} missing values"
                    )
            except Exception as e:
                self.errors.append(f"Error validating '{col}': {str(e)}")
        
        return len(self.errors) == 0
    
    def analyze(self):
        """
        Compute paper-compliant metrics for each participant.
        Returns list of dicts with: CUI, CDI, CCI, PSI, EI, LTMI, TI, NTI, CMI_P, expertise
        """
        if not self.validate():
            return None
        
        results = []
        
        for idx, row in self.df.iterrows():
            try:
                # Extract base indices
                cui = float(row['CUI'])
                cdi = float(row['CDI'])
                cci = float(row['CCI'])
                psi = float(row['PSI'])
                ei = float(row['EI'])
                ltmi = float(row['LTMI'])
                
                # Compute Technical Index (Bloom's weights: 1/6, 2/6, 3/6)
                # STEP 1: Apply Bloom's weights to each index
                u_s1 = cui * (0.6)
                d_s1 = cdi * (0.2)
                c_s1 = cci * (0.2)
                
                # STEP 2: Scale each weighted value from [-0.25, 1.0] to [0, 1]
                def scale_val(x):
                    return np.clip((x + 0.25) / 1.25, 0, 1)
                
                t1 = scale_val(u_s1)
                t2 = scale_val(d_s1)
                t3 = scale_val(c_s1)
                
                # STEP 3: Compute Technical Index as average of scaled values
                ti = (t1 + t2 + t3) / 3
                
                # Compute Non-Technical Index (equal average)
                nti = (psi + ei + ltmi) / 3
                
                # Compute Comprehension Measure Index (equal weighting)
                cmi_p = 0.5 * ti + 0.5 * nti
                
                # Classify expertise level
                if cmi_p >= 0.81:
                    expertise = 'Very High'
                elif cmi_p >= 0.61:
                    expertise = 'High'
                elif cmi_p >= 0.41:
                    expertise = 'Average'
                elif cmi_p >= 0.21:
                    expertise = 'Low'
                else:
                    expertise = 'Very Low'
                
                results.append({
                    'participant_idx': idx,
                    'CUI': cui,
                    'CDI': cdi,
                    'CCI': cci,
                    'PSI': psi,
                    'EI': ei,
                    'LTMI': ltmi,
                    'TI': ti,
                    'NTI': nti,
                    'CMI_P': cmi_p,
                    'expertise': expertise
                })
            except Exception as e:
                self.warnings.append(f"Row {idx}: {str(e)}")
        
        return results
    
    def get_summary_stats(self):
        """Compute summary statistics for all participants"""
        results = self.analyze()
        if not results:
            return None
        
        df_results = pd.DataFrame(results)
        
        return {
            'total_participants': len(results),
            'CUI_mean': float(df_results['CUI'].mean()),
            'CUI_std': float(df_results['CUI'].std()),
            'CDI_mean': float(df_results['CDI'].mean()),
            'CDI_std': float(df_results['CDI'].std()),
            'CCI_mean': float(df_results['CCI'].mean()),
            'CCI_std': float(df_results['CCI'].std()),
            'PSI_mean': float(df_results['PSI'].mean()),
            'PSI_std': float(df_results['PSI'].std()),
            'EI_mean': float(df_results['EI'].mean()),
            'EI_std': float(df_results['EI'].std()),
            'LTMI_mean': float(df_results['LTMI'].mean()),
            'LTMI_std': float(df_results['LTMI'].std()),
            'TI_mean': float(df_results['TI'].mean()),
            'TI_std': float(df_results['TI'].std()),
            'NTI_mean': float(df_results['NTI'].mean()),
            'NTI_std': float(df_results['NTI'].std()),
            'CMI_P_mean': float(df_results['CMI_P'].mean()),
            'CMI_P_std': float(df_results['CMI_P'].std()),
            'expertise_distribution': df_results['expertise'].value_counts().to_dict()
        }





# ==================== TECHNICAL METRICS COMPUTATION ====================

def compute_technical_from_normalized(df):
    """
    Compute comprehensive metrics from normalized Excel columns (technical & non-technical).
    
    INPUT: DataFrame with 6 normalized columns:
      Technical: U_S, D_S, C_S (domain: [-0.25, 1.0])
      Non-Technical: PSI, EI, LTMI (domain: [0, 1])
    
    OUTPUT: Enriched DataFrame with computed indices:
      Core indices: CUI, CDI, CCI
      Bloom-weighted: U_S1, D_S1, C_S1
      Scaled technical: T1, T2, T3
      Composite: TI, NTI, CMI_P
    
    Steps:
    1. Validate all 6 required columns exist
    2. Map technical to core indices: CUI = U_S, CDI = D_S, CCI = C_S
    3. Apply Bloom's weights: U_S1 = CUI × 0.6, D_S1 = CDI × 0.2, C_S1 = CCI × 0.2
    4. Scale to [0,1]: scale(x) = clip((x + 0.25) / 1.25, 0, 1) → T1, T2, T3
    5. Compute Technical Index: TI = (T1 + T2 + T3) / 3
    6. Compute Non-Technical Index: NTI = (PSI + EI + LTMI) / 3
    7. Compute Comprehension Measure: CMI_P = 0.5 × TI + 0.5 × NTI
    """
    
    # Work on a copy to avoid modifying original
    df_enriched = df.copy()
    
    # Validate all required columns
    technical_cols = ['U_S', 'D_S', 'C_S']
    non_technical_cols = ['PSI', 'EI', 'LTMI']
    required_cols = technical_cols + non_technical_cols
    
    missing = [col for col in required_cols if col not in df_enriched.columns]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")
    
    # ==================== TECHNICAL INDICES ====================
    
    # STEP 1: Map technical normalized to core indices
    df_enriched['CUI'] = df_enriched['U_S']
    df_enriched['CDI'] = df_enriched['D_S']
    df_enriched['CCI'] = df_enriched['C_S']
    
    # STEP 2: Apply Bloom's taxonomy weighting (technical only)
    df_enriched['U_S1'] = df_enriched['CUI'] * (0.6)
    df_enriched['D_S1'] = df_enriched['CDI'] * (0.2)
    df_enriched['C_S1'] = df_enriched['CCI'] * (0.2)
    
    # STEP 3: Scale Bloom-weighted values to [0,1] range
    def scale_to_unit(x):
        """Scale from [-0.25, 1.0] to [0, 1]: (x + 0.25) / 1.25, clipped to [0, 1]"""
        return np.clip((x + 0.25) / 1.25, 0, 1)
    
    df_enriched['T1'] = df_enriched['U_S1'].apply(scale_to_unit)
    df_enriched['T2'] = df_enriched['D_S1'].apply(scale_to_unit)
    df_enriched['T3'] = df_enriched['C_S1'].apply(scale_to_unit)
    
    # STEP 4: Compute Technical Index (TI)
    df_enriched['TI'] = (df_enriched['T1'] + df_enriched['T2'] + df_enriched['T3']) / 3
    
    # ==================== NON-TECHNICAL INDICES ====================
    
    # STEP 5: Compute Non-Technical Index (NTI) from existing PSI, EI, LTMI
    # NOTE: PSI, EI, LTMI are NOT modified, only averaged
    df_enriched['NTI'] = (df_enriched['PSI'] + df_enriched['EI'] + df_enriched['LTMI']) / 3
    
    # ==================== COMPREHENSION MEASURE ====================
    
    # STEP 6: Compute Comprehension Measure Index (CMI_P)
    df_enriched['CMI_P'] = 0.5 * df_enriched['TI'] + 0.5 * df_enriched['NTI']
    
    return df_enriched


# ==================== DATABASE INITIALIZATION ====================

def init_db():
    """Initialize database"""
    with app.app_context():
        db.create_all()
        
        admin = Admin.query.filter_by(username='admin').first()
        if not admin:
            admin = Admin(
                username='admin',
                email='admin@research.local',
                full_name='Research Administrator',
                role='super_admin'
            )
            admin.set_password('admin@2024')
            db.session.add(admin)
            db.session.commit()
            print("✓ Default admin created: username=admin, password=admin@2024")


# ==================== DATA LOADING HELPERS ====================

def get_dataframe_from_file(file_id_or_path):
    """Helper to get dataframe from file ID or path"""
    if isinstance(file_id_or_path, int):
        data_file = DataFile.query.get(file_id_or_path)
        if not data_file:
            return None
        file_path = data_file.file_path
    else:
        file_path = file_id_or_path
    
    try:
        if file_path.endswith('.csv'):
            df = pd.read_csv(file_path)
        else:
            df = pd.read_excel(file_path)
        return df
    except Exception as e:
        print(f"Error reading file: {e}")
        return None


# ==================== AUTHENTICATION ROUTES ====================

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('user_type') == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('user_dashboard'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Unified login page"""
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user_type = request.form.get('user_type', 'user')
        
        if user_type == 'admin':
            admin = Admin.query.filter_by(username=username).first()
            if admin and admin.check_password(password) and admin.is_active:
                session['user_id'] = admin.id
                session['username'] = admin.username
                session['user_type'] = 'admin'
                admin.last_login = datetime.utcnow()
                db.session.commit()
                flash(f'Welcome {admin.full_name}!', 'success')
                return redirect(url_for('admin_dashboard'))
            else:
                flash('Invalid admin credentials', 'danger')
        else:
            user = User.query.filter_by(username=username).first()
            if user and user.check_password(password) and user.is_active:
                session['user_id'] = user.id
                session['username'] = user.username
                session['user_type'] = 'user'
                user.last_login = datetime.utcnow()
                db.session.commit()
                flash(f'Welcome {user.full_name}!', 'success')
                return redirect(url_for('user_dashboard'))
            else:
                flash('Invalid user credentials', 'danger')
    
    return render_template('login.html')


@app.route('/register', methods=['GET', 'POST'])
def register():
    """User registration with comprehensive validation"""
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        email = request.form.get('email', '').strip()
        password = request.form.get('password', '')
        full_name = request.form.get('full_name', '').strip()
        
        # ==================== VALIDATION ====================
        
        # Check all fields are provided
        if not all([username, email, password, full_name]):
            flash('All fields are required', 'danger')
            return redirect(url_for('register'))
        
        # ==================== FULL NAME VALIDATION ====================
        if len(full_name) < 2:
            flash('Full name must be at least 2 characters', 'danger')
            return redirect(url_for('register'))
        
        if len(full_name) > 100:
            flash('Full name cannot exceed 100 characters', 'danger')
            return redirect(url_for('register'))
        
        # Check full name contains only valid characters
        if not re.match(r"^[a-zA-Z\s'-]{2,100}$", full_name):
            flash('Full name can only contain letters, spaces, hyphens, and apostrophes', 'danger')
            return redirect(url_for('register'))
        
        # ==================== USERNAME VALIDATION ====================
        if len(username) < 3:
            flash('Username must be at least 3 characters', 'danger')
            return redirect(url_for('register'))
        
        if len(username) > 20:
            flash('Username cannot exceed 20 characters', 'danger')
            return redirect(url_for('register'))
        
        # Check username contains only valid characters
        if not re.match(r"^[a-zA-Z0-9_-]{3,20}$", username):
            flash('Username can only contain letters, numbers, underscores, and hyphens', 'danger')
            return redirect(url_for('register'))
        
        # Check username uniqueness
        if User.query.filter_by(username=username).first():
            flash('Username already exists. Please choose a different username', 'danger')
            return redirect(url_for('register'))
        
        # ==================== EMAIL VALIDATION ====================
        if not re.match(r"^[^\s@]+@[^\s@]+\.[^\s@]+$", email):
            flash('Please enter a valid email address', 'danger')
            return redirect(url_for('register'))
        
        if len(email) > 120:
            flash('Email address is too long', 'danger')
            return redirect(url_for('register'))
        
        # Check email uniqueness
        if User.query.filter_by(email=email).first():
            flash('This email address is already registered. Please use a different email or login', 'danger')
            return redirect(url_for('register'))
        
        # ==================== PASSWORD VALIDATION ====================
        if len(password) < 8:
            flash('Password must be at least 8 characters', 'danger')
            return redirect(url_for('register'))
        
        if len(password) > 128:
            flash('Password is too long', 'danger')
            return redirect(url_for('register'))
        
        # Check password contains uppercase letter
        if not re.search(r"[A-Z]", password):
            flash('Password must contain at least one uppercase letter', 'danger')
            return redirect(url_for('register'))
        
        # Check password contains number
        if not re.search(r"\d", password):
            flash('Password must contain at least one number', 'danger')
            return redirect(url_for('register'))
        
        # Check password contains special character
        if not re.search(r"[!@#$%^&*()_+\-=\[\]{};':\"\\|,.<>\/?]", password):
            flash('Password must contain at least one special character (!@#$%^&*)', 'danger')
            return redirect(url_for('register'))
        
        # Check password is not too common
        common_passwords = ['Password1!', 'Welcome1!', 'Admin1234!', 'Test1234!']
        if password in common_passwords:
            flash('This password is too common. Please choose a stronger password', 'danger')
            return redirect(url_for('register'))
        
        # ==================== USER CREATION ====================
        try:
            user = User(username=username, email=email, full_name=full_name)
            user.set_password(password)
            db.session.add(user)
            db.session.commit()
            
            flash('Registration successful! Please login with your credentials.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash('An error occurred during registration. Please try again.', 'danger')
            return redirect(url_for('register'))
    
    return render_template('register.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('login'))


# ==================== ADMIN ROUTES ====================

@app.route('/admin/dashboard')
def admin_dashboard():
    """Admin dashboard"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    admin = Admin.query.get(session['user_id'])
    files_count = DataFile.query.filter_by(is_active=True).count()
    users_count = User.query.filter_by(is_active=True).count()
    recent_files = DataFile.query.filter_by(is_active=True).order_by(DataFile.uploaded_at.desc()).limit(5).all()
    
    stats = {
        'files': files_count,
        'users': users_count,
        'last_upload': recent_files[0].uploaded_at if recent_files else None
    }
    
    return render_template('admin/dashboard.html', admin=admin, stats=stats, files=recent_files)


@app.route('/admin/upload', methods=['GET', 'POST'])
def admin_upload():
    """File upload"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        description = request.form.get('description', '')
        
        if file.filename == '' or not allowed_file(file.filename):
            flash('Invalid file', 'danger')
            return redirect(request.url)
        
        try:
            if file.filename.endswith('.csv'):
                df = pd.read_csv(file)
            else:
                df = pd.read_excel(file)
            
            rows, cols = df.shape
            file.seek(0)
            
            # Ensure only one dataset stays active for users
            existing_active_files = DataFile.query.filter_by(is_active=True).all()
            for existing_file in existing_active_files:
                existing_file.is_active = False
            
            timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S_')
            filename = timestamp + secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            data_file = DataFile(
                filename=filename,
                original_name=file.filename,
                file_path=filepath,
                uploaded_by=session['username'],
                file_size=os.path.getsize(filepath),
                rows_count=rows,
                columns_count=cols,
                description=description
            )
            db.session.add(data_file)
            db.session.commit()
            
            flash(f'✓ File uploaded! ({rows} rows, {cols} columns)', 'success')
            return redirect(url_for('admin_data_manager'))
        
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('admin/upload.html')


@app.route('/admin/data-manager')
def admin_data_manager():
    """Data manager"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    active_file = DataFile.query.filter_by(is_active=True).order_by(DataFile.uploaded_at.desc()).first()
    archived_files = DataFile.query.filter_by(is_active=False).order_by(DataFile.uploaded_at.desc()).all()
    return render_template('admin/data_manager.html', active_file=active_file, archived_files=archived_files)


@app.route('/admin/data-file/<int:file_id>/archive', methods=['POST'])
def admin_archive_file(file_id):
    """Archive (remove) current dataset from user view"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    data_file = DataFile.query.get(file_id)
    if not data_file:
        flash('File not found', 'danger')
        return redirect(url_for('admin_data_manager'))
    
    data_file.is_active = False
    db.session.commit()
    flash('Previous upload archived. Users will no longer see this dataset.', 'success')
    return redirect(url_for('admin_data_manager'))


@app.route('/admin/data-file/<int:file_id>/activate', methods=['POST'])
def admin_activate_file(file_id):
    """Set selected dataset as the only active file"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    data_file = DataFile.query.get(file_id)
    if not data_file:
        flash('File not found', 'danger')
        return redirect(url_for('admin_data_manager'))
    
    for file_record in DataFile.query.all():
        file_record.is_active = (file_record.id == file_id)
    db.session.commit()
    flash('Dataset activated and is now visible to users.', 'success')
    return redirect(url_for('admin_data_manager'))


@app.route('/admin/data-file/<int:file_id>/download')
def admin_download_file(file_id):
    """Download stored dataset regardless of status"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    data_file = DataFile.query.get(file_id)
    if not data_file or not os.path.exists(data_file.file_path):
        flash('File not found on server', 'danger')
        return redirect(url_for('admin_data_manager'))
    
    return send_file(data_file.file_path, as_attachment=True, download_name=data_file.original_name)


@app.route('/admin/configuration')
def admin_configuration():
    """Configuration"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    return render_template('admin/configuration.html')


@app.route('/admin/single-analysis')
def admin_single_analysis():
    """Single student analysis"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    latest = DataFile.query.filter_by(is_active=True).order_by(DataFile.uploaded_at.desc()).first()
    return render_template('admin/single_analysis.html', latest=latest)


@app.route('/admin/quick-analyze/<int:file_id>', methods=['GET', 'POST'])
def admin_quick_analyze(file_id):
    """Quick analysis creation using paper-compliant metrics"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    data_file = DataFile.query.get(file_id)
    if not data_file:
        flash('File not found', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    if request.method == 'POST':
        title = request.form.get('title', f'Analysis - {data_file.original_name}')
        description = request.form.get('description', '')
        
        try:
            # Read file
            df = get_dataframe_from_file(file_id)
            if df is None:
                flash('Error reading file', 'danger')
                return redirect(url_for('admin_dashboard'))
            
            # Use paper-compliant analyzer
            analyzer = PaperCompliantAnalyzer(df)
            if not analyzer.validate():
                errors_str = '; '.join(analyzer.errors)
                flash(f'❌ Validation failed: {errors_str}', 'danger')
                return redirect(url_for('admin_quick_analyze', file_id=file_id))
            
            # Get summary statistics
            stats = analyzer.get_summary_stats()
            
            # Create metrics dict
            metrics = {
                'summary_stats': stats,
                'participants': analyzer.analyze(),
                'metadata': {
                    'analyzer_type': 'paper_compliant',
                    'analysis_date': datetime.utcnow().isoformat(),
                    'note': 'Analysis based on paper-compliant metrics (TI, NTI, CMI_P)'
                }
            }
            
            # Create analysis record
            analysis = Analysis(
                title=title,
                description=description,
                data_file_id=file_id,
                created_by=session['username'],
                metrics_json=json.dumps(metrics)
            )
            db.session.add(analysis)
            db.session.commit()
            
            flash(f'✓ Analysis created successfully! ({stats["total_participants"]} participants)', 'success')
            return redirect(url_for('admin_view_analysis', analysis_id=analysis.id))
        
        except Exception as e:
            flash(f'Error during analysis: {str(e)}', 'danger')
            return redirect(url_for('admin_data_manager'))
    
    return render_template('admin/quick_analyze.html', file=data_file)


@app.route('/admin/bulk-analysis')
def admin_bulk_analysis():
    """Bulk analysis"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    latest = DataFile.query.filter_by(is_active=True).order_by(DataFile.uploaded_at.desc()).first()
    return render_template('admin/bulk_analysis.html', latest=latest)


@app.route('/admin/advanced-calc')
def admin_advanced_calc():
    """Interactive Performance Calculator"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    # Get list of all active users
    users = User.query.filter_by(is_active=True).order_by(User.full_name).all()
    
    return render_template('admin/calculator.html', users=users)


@app.route('/admin/download-calculation', methods=['POST'])
def download_calculation():
    """Download performance calculation results as Excel file"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
    
    try:
        import json
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        # Get calculation data
        calc_data = json.loads(request.form.get('data', '{}'))
        selected_user_id = calc_data.get('selected_user_id')
        
        # Get selected user info if provided
        selected_user_info = None
        if selected_user_id:
            try:
                selected_user = User.query.get(int(selected_user_id))
                if selected_user:
                    selected_user_info = {
                        'id': selected_user.id,
                        'name': selected_user.full_name,
                        'email': selected_user.email
                    }
            except Exception as e:
                app.logger.error(f"Error fetching selected user: {str(e)}")
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Performance Results"
        
        # Set up styles
        title_font = Font(bold=True, size=14, color="FFFFFF")
        title_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
        
        section_font = Font(bold=True, size=11, color="FFFFFF")
        section_fill = PatternFill(start_color="1e40af", end_color="1e40af", fill_type="solid")
        
        sub_header_fill = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
        sub_header_font = Font(bold=True, size=10, color="1e40af")
        
        label_font = Font(bold=True, size=9, color="1e293b")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Title
        ws[f'A{row}'] = "COGNITIVE & EMOTIONAL PERFORMANCE RESULTS"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].fill = title_fill
        ws[f'A{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells(f'A{row}:D{row}')
        ws.row_dimensions[row].height = 25
        row += 1
        
        # Student Information Section
        student_info = calc_data.get('student', {})
        
        ws[f'A{row}'] = "STUDENT INFORMATION"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        student_data = [
            ('Student ID', student_info.get('id', 'N/A')),
            ('Student Name', student_info.get('name', 'N/A')),
            ('Assessment Date', student_info.get('date', 'N/A'))
        ]
        
        for idx, (label, value) in enumerate(student_data, 1):
            ws[f'A{row}'] = f"{idx}. {label}"
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = label_font
            ws[f'A{row}'].border = border
            ws[f'B{row}'].border = border
            ws.merge_cells(f'B{row}:D{row}')
            row += 1
        
        # Add selected user assignment section if user was selected
        if selected_user_info:
            row += 1
            ws[f'A{row}'] = "ASSIGNED TO"
            ws[f'A{row}'].font = section_font
            ws[f'A{row}'].fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            user_data = [
                ('User Name', selected_user_info.get('name', 'N/A')),
                ('User Email', selected_user_info.get('email', 'N/A'))
            ]
            
            for idx, (label, value) in enumerate(user_data, 1):
                ws[f'A{row}'] = f"{idx}. {label}"
                ws[f'B{row}'] = value
                ws[f'A{row}'].font = label_font
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                ws.merge_cells(f'B{row}:D{row}')
                row += 1
        
        row += 1
        
        # Input Data Section
        ws[f'A{row}'] = "INPUT DATA"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Technical inputs header
        inputs = calc_data.get('inputs', {})
        
        ws[f'A{row}'] = "#"
        ws[f'B{row}'] = "Domain"
        ws[f'C{row}'] = "Correct"
        ws[f'D{row}'] = "Wrong"
        for cell_ref in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
            ws[cell_ref].fill = sub_header_fill
            ws[cell_ref].font = sub_header_font
            ws[cell_ref].border = border
        row += 1
        
        item_num = 1
        for domain, data in inputs.items():
            if domain.startswith('emotion'):
                continue
            ws[f'A{row}'] = item_num
            ws[f'B{row}'] = domain.capitalize()
            ws[f'C{row}'] = data.get('correct', 0)
            ws[f'D{row}'] = data.get('wrong', 0)
            for cell_ref in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                ws[cell_ref].border = border
            item_num += 1
            row += 1
        
        row += 1
        
        # Calculated Results Section
        ws[f'A{row}'] = "CALCULATED RESULTS"
        ws[f'A{row}'].font = section_font
        ws[f'A{row}'].fill = section_fill
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        # Technical scores
        ws[f'A{row}'] = "TECHNICAL PERFORMANCE"
        ws[f'A{row}'].font = Font(bold=True, color="667eea", size=11)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        technical = calc_data.get('technical', {})
        tech_scores = [
            ('Understanding (T1)', technical.get('t1', 0), 'Score out of 1.0'),
            ('Debugging (T2)', technical.get('t2', 0), 'Score out of 1.0'),
            ('Completion (T3)', technical.get('t3', 0), 'Score out of 1.0'),
            ('Technical Total (CL1_Div3_Scaled)', technical.get('cl1_div3_scaled', 0), 'Score out of 0.5')
        ]
        
        tech_num = 1
        for label, value, desc in tech_scores:
            ws[f'A{row}'] = f"{tech_num}."
            ws[f'B{row}'] = label
            ws[f'C{row}'] = round(value, 4)
            ws[f'D{row}'] = desc
            for cell_ref in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                ws[cell_ref].border = border
            tech_num += 1
            row += 1
        
        row += 1
        
        # Emotional scores
        ws[f'A{row}'] = "EMOTIONAL PERFORMANCE"
        ws[f'A{row}'].font = Font(bold=True, color="f5576c", size=11)
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        
        emotional = calc_data.get('emotional', {})
        active_emotion_count = emotional.get('activeEmotionCount', 7)
        
        emotion_scores = [
            ('Before Average (B1)', emotional.get('b1', 0), 'Average before assessment'),
            ('After Average (A1)', emotional.get('a1', 0), 'Average after assessment'),
            ('Emotional Total (Emotion_Value_Scaled)', emotional.get('emotion_value_scaled', 0), f'Score out of 0.25 (based on {active_emotion_count} parameters)')
        ]
        
        emotion_num = 1
        for label, value, desc in emotion_scores:
            ws[f'A{row}'] = f"{emotion_num}."
            ws[f'B{row}'] = label
            ws[f'C{row}'] = round(value, 4)
            ws[f'D{row}'] = desc
            for cell_ref in [f'A{row}', f'B{row}', f'C{row}', f'D{row}']:
                ws[cell_ref].border = border
            emotion_num += 1
            row += 1
        
        # Show active emotions used
        row += 1
        active_emotions = calc_data.get('inputs', {}).get('active_emotions', [])
        if active_emotions:
            ws[f'A{row}'] = "ACTIVE EMOTION PARAMETERS"
            ws[f'A{row}'].font = Font(bold=True, size=10, color="FFFFFF")
            ws[f'A{row}'].fill = PatternFill(start_color="8b5cf6", end_color="8b5cf6", fill_type="solid")
            ws.merge_cells(f'A{row}:D{row}')
            row += 1
            
            for idx, emotion in enumerate(active_emotions, 1):
                ws[f'A{row}'] = f"{idx}."
                ws[f'B{row}'] = emotion
                ws[f'A{row}'].border = border
                ws[f'B{row}'].border = border
                ws.merge_cells(f'B{row}:D{row}')
                row += 1
        
        row += 1
        
        # Combined score
        combined = calc_data.get('combined', 0)
        percentage = calc_data.get('percentage', 0)
        
        ws[f'A{row}'] = "COMBINED PERFORMANCE SCORE"
        ws[f'A{row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid")
        ws.merge_cells(f'A{row}:D{row}')
        row += 1
        ws[f'B{row}'] = round(combined, 4)
        ws[f'C{row}'] = "out of 0.75"
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        ws[f'B{row}'].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        ws[f'C{row}'].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        for cell_ref in [f'A{row}', f'B{row}', f'C{row}']:
            ws[cell_ref].border = border
        row += 1
        
        ws[f'A{row}'] = "Performance %"
        ws[f'B{row}'] = f"{round(percentage, 1)}%"
        ws[f'C{row}'] = "out of 100%"
        ws[f'A{row}'].font = label_font
        ws[f'A{row}'].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        ws[f'B{row}'].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        ws[f'C{row}'].fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        for cell_ref in [f'A{row}', f'B{row}', f'C{row}']:
            ws[cell_ref].border = border
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 4
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 25
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        student_id = student_info.get('id', 'Unknown')
        filename = f"Performance_{student_id}_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        flash(f'Error generating Excel: {str(e)}', 'danger')
        return redirect(url_for('admin_advanced_calc'))


@app.route('/admin/technical-upload', methods=['GET', 'POST'])
def admin_technical_upload():
    """Upload and compute Bloom's taxonomy-based technical metrics"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        if 'file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        file = request.files['file']
        title = request.form.get('title', 'Technical Analysis')
        description = request.form.get('description', '')
        
        if file.filename == '' or not allowed_file(file.filename):
            flash('Invalid file format (use .csv or .xlsx)', 'danger')
            return redirect(request.url)
        
        try:
            # Read original file (keeping it unchanged)
            if file.filename.endswith('.csv'):
                df_original = pd.read_csv(file)
            else:
                df_original = pd.read_excel(file)
            
            # Compute technical metrics from normalized columns
            df_enriched = compute_technical_from_normalized(df_original)
            
            # Save original uploaded file unchanged
            file.seek(0)
            timestamp = datetime.utcnow().strftime('%Y%m%d_%H%M%S_')
            filename = timestamp + secure_filename(file.filename)
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(filepath)
            
            # Save enriched dataset with computed metrics
            enriched_filename = f"enriched_{timestamp}{file.filename.rsplit('.', 1)[0]}.xlsx"
            enriched_path = os.path.join(app.config['UPLOAD_FOLDER'], enriched_filename)
            df_enriched.to_excel(enriched_path, index=False)
            
            # Prepare summary statistics for visualization
            stats = {
                'total_participants': len(df_enriched),
                'CUI_mean': float(df_enriched['CUI'].mean()),
                'CUI_std': float(df_enriched['CUI'].std()),
                'CDI_mean': float(df_enriched['CDI'].mean()),
                'CDI_std': float(df_enriched['CDI'].std()),
                'CCI_mean': float(df_enriched['CCI'].mean()),
                'CCI_std': float(df_enriched['CCI'].std()),
                'T1_mean': float(df_enriched['T1'].mean()),
                'T1_std': float(df_enriched['T1'].std()),
                'T2_mean': float(df_enriched['T2'].mean()),
                'T2_std': float(df_enriched['T2'].std()),
                'T3_mean': float(df_enriched['T3'].mean()),
                'T3_std': float(df_enriched['T3'].std()),
                'PSI_mean': float(df_enriched['PSI'].mean()),
                'PSI_std': float(df_enriched['PSI'].std()),
                'EI_mean': float(df_enriched['EI'].mean()),
                'EI_std': float(df_enriched['EI'].std()),
                'LTMI_mean': float(df_enriched['LTMI'].mean()),
                'LTMI_std': float(df_enriched['LTMI'].std()),
                'TI_mean': float(df_enriched['TI'].mean()),
                'TI_std': float(df_enriched['TI'].std()),
                'NTI_mean': float(df_enriched['NTI'].mean()),
                'NTI_std': float(df_enriched['NTI'].std()),
                'CMI_P_mean': float(df_enriched['CMI_P'].mean()),
                'CMI_P_std': float(df_enriched['CMI_P'].std())
            }
            
            # Create record in database
            tech_analysis = TechnicalAnalysis(
                title=title,
                description=description,
                uploaded_file_path=filepath,
                original_filename=file.filename,
                uploaded_by=session['username'],
                rows_count=len(df_enriched),
                enriched_file_path=enriched_path,
                viz_data_json=json.dumps(stats),
                validation_status='valid'
            )
            db.session.add(tech_analysis)
            db.session.commit()
            
            flash(f'✓ Comprehensive analysis computed successfully! ({len(df_enriched)} participants) - Technical & Non-Technical indices', 'success')
            return redirect(url_for('admin_technical_dashboard', analysis_id=tech_analysis.id))
        
        except ValueError as e:
            flash(f'❌ Validation error: {str(e)}', 'danger')
            return redirect(request.url)
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(request.url)
    
    return render_template('admin/technical_upload.html')


@app.route('/admin/technical-dashboard/<int:analysis_id>')
def admin_technical_dashboard(analysis_id):
    """Technical analysis dashboard with visualizations"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    analysis = TechnicalAnalysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_advanced_calc'))
    
    stats = json.loads(analysis.viz_data_json) if analysis.viz_data_json else {}
    
    return render_template('admin/technical_dashboard_v2.html', analysis=analysis, stats=stats)


@app.route('/admin/technical-download/<int:analysis_id>')
def admin_technical_download(analysis_id):
    """Download enriched technical dataset"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    analysis = TechnicalAnalysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_advanced_calc'))
    
    # Convert relative paths to absolute
    enriched_path = analysis.enriched_file_path
    if not os.path.isabs(enriched_path):
        enriched_path = os.path.abspath(enriched_path)
    
    if not os.path.exists(enriched_path):
        flash('Enriched file not found', 'danger')
        return redirect(url_for('admin_advanced_calc'))
    
    return send_file(
        enriched_path,
        as_attachment=True,
        download_name=f"enriched_technical_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    )


# ==================== API ROUTES FOR TECHNICAL ANALYSIS ====================

@app.route('/api/technical/chart-data/<int:analysis_id>')
def api_technical_chart_data(analysis_id):
    """Get chart data for technical analysis including per-student data"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    analysis = TechnicalAnalysis.query.get(analysis_id)
    if not analysis:
        return jsonify({'error': 'Analysis not found'}), 404
    
    stats = json.loads(analysis.viz_data_json) if analysis.viz_data_json else {}
    
    try:
        # Convert relative paths to absolute
        enriched_path = analysis.enriched_file_path
        if not os.path.isabs(enriched_path):
            enriched_path = os.path.abspath(enriched_path)
        
        if not os.path.exists(enriched_path):
            return jsonify({'error': f'Enriched file not found: {enriched_path}'}), 404
        
        # Load enriched dataset for per-student analysis
        df_enriched = pd.read_excel(enriched_path)
        
        # Prepare student-level data
        students = []
        for idx, row in df_enriched.iterrows():
            student_id = str(row.get('Student_ID', f'STU{idx+1:03d}'))
            students.append({
                'id': student_id,
                'idx': idx,
                'CUI': float(row['CUI']),
                'CDI': float(row['CDI']),
                'CCI': float(row['CCI']),
                'T1': float(row['T1']),
                'T2': float(row['T2']),
                'T3': float(row['T3']),
                'CL1': float(row['CL1']),
                'CL1_Div3': float(row.get('CL1_Div3', row['CL1'] / 3)),
                'CL1_Div3_Scaled': float(row['CL1_Div3_Scaled'])
            })
        
        return jsonify({
            'stats': stats,
            'student_count': len(df_enriched),
            'students': students,
            'cui_values': df_enriched['CUI'].tolist(),
            'cdi_values': df_enriched['CDI'].tolist(),
            'cci_values': df_enriched['CCI'].tolist(),
            't1_values': df_enriched['T1'].tolist(),
            't2_values': df_enriched['T2'].tolist(),
            't3_values': df_enriched['T3'].tolist(),
            'cl1_values': df_enriched['CL1'].tolist(),
            'cl1_div3_scaled': df_enriched['CL1_Div3_Scaled'].tolist()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/admin/analyze', methods=['GET', 'POST'])
def admin_analyze():
    """Admin analyzes uploaded data using paper-compliant metrics"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        file_id = request.form.get('file_id')
        title = request.form.get('title')
        description = request.form.get('description')
        
        data_file = DataFile.query.get(file_id)
        if not data_file:
            flash('File not found', 'danger')
            return redirect(url_for('admin_dashboard'))
        
        try:
            # Read file
            df = get_dataframe_from_file(file_id)
            if df is None:
                flash('Error reading file', 'danger')
                return redirect(url_for('admin_dashboard'))
            
            # Use paper-compliant analyzer
            analyzer = PaperCompliantAnalyzer(df)
            if not analyzer.validate():
                errors_str = '; '.join(analyzer.errors)
                flash(f'❌ Validation failed: {errors_str}', 'danger')
                return redirect(url_for('admin_dashboard'))
            
            # Show any warnings
            for warning in analyzer.warnings:
                flash(f'⚠️ {warning}', 'warning')
            
            # Get summary statistics
            stats = analyzer.get_summary_stats()
            
            # Create metrics dict for storage
            metrics = {
                'summary_stats': stats,
                'participants': analyzer.analyze(),
                'metadata': {
                    'analyzer_type': 'paper_compliant',
                    'analysis_date': datetime.utcnow().isoformat(),
                    'note': 'Analysis based on paper-compliant metrics (TI, NTI, CMI_P)'
                }
            }
            
            # Create analysis record
            analysis = Analysis(
                title=title,
                description=description,
                data_file_id=file_id,
                created_by=session['username'],
                metrics_json=json.dumps(metrics)
            )
            db.session.add(analysis)
            db.session.commit()
            
            flash(f'✓ Analysis created successfully! ({stats["total_participants"]} participants)', 'success')
            return redirect(url_for('admin_view_analysis', analysis_id=analysis.id))
        
        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')
            return redirect(url_for('admin_dashboard'))
    
    files = DataFile.query.filter_by(is_active=True).all()
    return render_template('admin/analyze.html', files=files)


@app.route('/admin/analysis/<int:analysis_id>')
def admin_view_analysis(analysis_id):
    """View analysis results"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    analysis = Analysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    # Fetch the data_file
    if analysis.data_file_id:
        analysis.data_file = DataFile.query.get(analysis.data_file_id)
    
    metrics = json.loads(analysis.metrics_json) if analysis.metrics_json else {}
    
    # Get all registered users for sharing (include both active and inactive for admin visibility)
    users = User.query.order_by(User.full_name).all()
    
    return render_template('admin/view_analysis_paper.html', analysis=analysis, metrics=metrics, users=users)


@app.route('/admin/publish/<int:analysis_id>', methods=['POST'])
def admin_publish_analysis(analysis_id):
    """Publish analysis for users to see"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        flash('Unauthorized', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    analysis = Analysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_analyses'))
    
    # Check if already published
    if analysis.is_published:
        flash('This analysis is already published!', 'info')
        return redirect(url_for('admin_view_analysis', analysis_id=analysis_id))
    
    try:
        # Optional: unpublish older analyses from the same data file
        replace_old = request.form.get('replace_old') == 'on'
        
        if replace_old:
            # Find other published analyses from the same data file
            old_analyses = Analysis.query.filter(
                Analysis.data_file_id == analysis.data_file_id,
                Analysis.id != analysis.id,
                Analysis.is_published == True
            ).all()
            
            # Unpublish and remove permissions from old analyses
            for old_analysis in old_analyses:
                old_analysis.is_published = False
                # Remove all permissions for old analysis
                UserAnalysisPermission.query.filter_by(analysis_id=old_analysis.id).delete()
        
        # Fetch the data_file separately to ensure it's loaded
        if analysis.data_file_id:
            analysis.data_file = DataFile.query.get(analysis.data_file_id)
        
        analysis.is_published = True
        analysis.published_at = datetime.utcnow()
        db.session.commit()
        
        if replace_old:
            flash(f'✓ Analysis published! Older analyses from this dataset have been unpublished.', 'success')
        else:
            flash('✓ Analysis published! Users can now view it.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error publishing analysis: {str(e)}', 'danger')
    
    return redirect(url_for('admin_view_analysis', analysis_id=analysis_id))


@app.route('/admin/delete-analysis/<int:analysis_id>', methods=['GET', 'POST'])
def admin_delete_analysis(analysis_id):
    """Delete analysis and remove all shared permissions"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        flash('Unauthorized', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    analysis = Analysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_analyses'))
    
    try:
        # First, delete all permissions for this analysis
        UserAnalysisPermission.query.filter_by(analysis_id=analysis_id).delete()
        
        # Then delete the analysis itself
        analysis_title = analysis.title
        db.session.delete(analysis)
        db.session.commit()
        
        flash(f'✓ Analysis "{analysis_title}" and all shared permissions deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting analysis: {str(e)}', 'danger')
    
    return redirect(url_for('admin_analyses'))


@app.route('/admin/unpublish/<int:analysis_id>', methods=['POST'])
def admin_unpublish_analysis(analysis_id):
    """Unpublish analysis (remove from user view)"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        flash('Unauthorized', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    analysis = Analysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_analyses'))
    
    analysis.is_published = False
    analysis.published_at = None
    db.session.commit()
    
    flash('Analysis unpublished (hidden from users)', 'success')
    return redirect(url_for('admin_view_analysis', analysis_id=analysis_id))


@app.route('/admin/share-analysis/<int:analysis_id>', methods=['POST'])
def admin_share_analysis(analysis_id):
    """Share analysis with multiple users"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        flash('Unauthorized', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    analysis = Analysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_analyses'))
    
    # Get selected user IDs (can be multiple)
    user_ids = request.form.getlist('user_ids')
    if not user_ids:
        flash('Please select at least one user', 'warning')
        return redirect(url_for('admin_view_analysis', analysis_id=analysis_id))
    
    try:
        # Verify all users exist and add permissions
        selected_users = []
        for user_id_str in user_ids:
            user = User.query.get(int(user_id_str))
            if not user:
                flash(f'User ID {user_id_str} not found', 'danger')
                return redirect(url_for('admin_view_analysis', analysis_id=analysis_id))
            
            # Check if permission already exists
            existing = UserAnalysisPermission.query.filter_by(
                user_id=user.id, 
                analysis_id=analysis.id
            ).first()
            
            if not existing:
                # Create new permission
                permission = UserAnalysisPermission(
                    user_id=user.id,
                    analysis_id=analysis.id,
                    shared_at=datetime.utcnow()
                )
                db.session.add(permission)
            
            selected_users.append(user.full_name)
        
        db.session.commit()
        user_names = ', '.join(selected_users)
        flash(f'✓ Analysis shared with {len(selected_users)} user(s): {user_names}', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error sharing analysis: {str(e)}', 'danger')
    
    return redirect(url_for('admin_view_analysis', analysis_id=analysis_id))


@app.route('/admin/unshare-analysis/<int:analysis_id>', methods=['POST'])
def admin_unshare_analysis(analysis_id):
    """Remove sharing for specific user(s)"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        flash('Unauthorized', 'danger')
        return redirect(url_for('admin_dashboard'))
    
    analysis = Analysis.query.get(analysis_id)
    if not analysis:
        flash('Analysis not found', 'danger')
        return redirect(url_for('admin_analyses'))
    
    # Get user IDs to unshare (optional - if none, remove all)
    user_ids = request.form.getlist('user_ids')
    
    try:
        if user_ids:
            # Remove permissions for specific users
            for user_id_str in user_ids:
                permission = UserAnalysisPermission.query.filter_by(
                    user_id=int(user_id_str),
                    analysis_id=analysis.id
                ).first()
                if permission:
                    db.session.delete(permission)
            flash('✓ Sharing removed for selected user(s)', 'success')
        else:
            # Remove all permissions for this analysis
            UserAnalysisPermission.query.filter_by(analysis_id=analysis.id).delete()
            flash('✓ Analysis unshared from all users', 'success')
        
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        flash(f'Error removing sharing: {str(e)}', 'danger')
    
    return redirect(url_for('admin_view_analysis', analysis_id=analysis_id))


@app.route('/admin/analyses')
def admin_analyses():
    """List all analyses"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    analyses = Analysis.query.order_by(Analysis.created_at.desc()).all()
    return render_template('admin/analyses.html', analyses=analyses)


# ==================== USER ROUTES ====================

@app.route('/dashboard')
def dashboard():
    """Redirect to appropriate dashboard based on user type"""
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('user_type') == 'admin':
        return redirect(url_for('admin_dashboard'))
    else:
        return redirect(url_for('user_dashboard'))


@app.route('/user/dashboard')
def user_dashboard():
    """User dashboard - view analyses shared with them"""
    if 'user_id' not in session or session.get('user_type') != 'user':
        return redirect(url_for('login'))
    
    user = User.query.get(session['user_id'])
    
    # Debug logging
    print(f"\n=== USER DASHBOARD DEBUG ===")
    print(f"Current user ID: {session['user_id']}")
    print(f"User full name: {user.full_name if user else 'NOT FOUND'}")
    
    # Get all analyses shared with this user
    shared_analyses = Analysis.query.join(
        UserAnalysisPermission,
        UserAnalysisPermission.analysis_id == Analysis.id
    ).filter(
        UserAnalysisPermission.user_id == session['user_id']
    ).order_by(UserAnalysisPermission.shared_at.desc()).all()
    
    print(f"Shared analyses count: {len(shared_analyses)}")
    for analysis in shared_analyses:
        print(f"  - Analysis ID: {analysis.id}, Title: {analysis.title}")
    
    # Check if there are any permissions at all
    all_perms = UserAnalysisPermission.query.filter_by(user_id=session['user_id']).all()
    print(f"Total permissions for this user: {len(all_perms)}")
    for perm in all_perms:
        print(f"  - Permission ID: {perm.id}, Analysis ID: {perm.analysis_id}")
    
    print(f"=== END DEBUG ===\n")
    
    # Calculate statistics
    total_participants = sum(a.data_file.rows_count for a in shared_analyses if a.data_file)
    research_areas = set(a.title.split('-')[0].strip() if '-' in a.title else a.title for a in shared_analyses)
    
    return render_template('user/dashboard.html', user=user, published_analyses=shared_analyses, total_participants=total_participants, research_areas=research_areas)


@app.route('/user/analysis/<int:analysis_id>')
def user_view_analysis(analysis_id):
    """User view analysis shared with them"""
    if 'user_id' not in session or session.get('user_type') != 'user':
        return redirect(url_for('login'))
    
    analysis = Analysis.query.get(analysis_id)
    
    # Check if analysis is shared with this user
    permission = UserAnalysisPermission.query.filter_by(
        user_id=session['user_id'],
        analysis_id=analysis_id
    ).first()
    
    if not analysis or not permission:
        flash('Analysis not found or not shared with you', 'danger')
        return redirect(url_for('user_dashboard'))
    
    # Fetch the data_file
    if analysis.data_file_id:
        analysis.data_file = DataFile.query.get(analysis.data_file_id)
    
    # Load metrics from stored JSON
    raw_metrics = json.loads(analysis.metrics_json) if analysis.metrics_json else {}
    
    # Create a proper metrics object with defaults
    summary_stats = raw_metrics.get('summary_stats', {})
    
    # Ensure all required keys exist with defaults
    metrics = {
        'summary_stats': {
            'total_participants': summary_stats.get('total_participants', 0),
            'CUI_mean': summary_stats.get('CUI_mean', 0),
            'CUI_std': summary_stats.get('CUI_std', 0),
            'CDI_mean': summary_stats.get('CDI_mean', 0),
            'CDI_std': summary_stats.get('CDI_std', 0),
            'CCI_mean': summary_stats.get('CCI_mean', 0),
            'CCI_std': summary_stats.get('CCI_std', 0),
            'PSI_mean': summary_stats.get('PSI_mean', 0),
            'PSI_std': summary_stats.get('PSI_std', 0),
            'EI_mean': summary_stats.get('EI_mean', 0),
            'EI_std': summary_stats.get('EI_std', 0),
            'LTMI_mean': summary_stats.get('LTMI_mean', 0),
            'LTMI_std': summary_stats.get('LTMI_std', 0),
            'TI_mean': summary_stats.get('TI_mean', 0),
            'TI_std': summary_stats.get('TI_std', 0),
            'NTI_mean': summary_stats.get('NTI_mean', 0),
            'NTI_std': summary_stats.get('NTI_std', 0),
            'CMI_P_mean': summary_stats.get('CMI_P_mean', 0),
            'CMI_P_std': summary_stats.get('CMI_P_std', 0),
            'expertise_distribution': summary_stats.get('expertise_distribution', {'Very High': 0, 'High': 0, 'Average': 0, 'Low': 0, 'Very Low': 0})
        },
        'participants': raw_metrics.get('participants', []),
        'metadata': raw_metrics.get('metadata', {})
    }
    
    return render_template('user/view_analysis.html', analysis=analysis, metrics=metrics)


@app.route('/profile')
def user_profile():
    """User profile page"""
    if 'user_id' not in session or session.get('user_type') != 'user':
        return redirect(url_for('login'))
    
    user = User.query.get(session['user_id'])
    published_analyses = Analysis.query.filter_by(is_published=True).order_by(Analysis.published_at.desc()).all()
    
    # Get statistics
    total_analyses = len(published_analyses)
    last_viewed = published_analyses[0].title if published_analyses else None
    
    return render_template('user/profile.html', user_info=user, published_analyses=published_analyses, total_analyses=total_analyses, last_viewed=last_viewed)


@app.route('/eeg-data')
def user_eeg_data():
    """User EEG data page - display shared analyses"""
    if 'user_id' not in session or session.get('user_type') != 'user':
        return redirect(url_for('login'))
    
    user = User.query.get(session['user_id'])
    
    # Get all analyses shared with this user (same as dashboard)
    shared_analyses = Analysis.query.join(
        UserAnalysisPermission,
        UserAnalysisPermission.analysis_id == Analysis.id
    ).filter(
        UserAnalysisPermission.user_id == session['user_id']
    ).order_by(UserAnalysisPermission.shared_at.desc()).all()
    
    # Get statistics
    total_eeg_datasets = len(shared_analyses)
    total_participants = sum(a.data_file.rows_count for a in shared_analyses if a.data_file)
    
    return render_template('user/eeg_data.html', user=user, eeg_analyses=shared_analyses, total_datasets=total_eeg_datasets, total_participants=total_participants)


@app.route('/admin/eeg-data')
def admin_eeg_data():
    """Admin EEG data management page"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    # Get all analyses (including published ones that can be shared)
    all_analyses = Analysis.query.order_by(Analysis.created_at.desc()).all()
    
    # Get all users (for sharing)
    all_users = User.query.filter_by(user_type='user').all()
    
    # Get all data files
    all_files = DataFile.query.order_by(DataFile.uploaded_at.desc()).all()
    
    return render_template('admin/eeg_data.html', analyses=all_analyses, users=all_users, data_files=all_files)


@app.route('/admin/user-management')
def admin_user_management():
    """Admin user data management page"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    # Get all users and admins
    all_users = User.query.all()
    all_admins = Admin.query.all()
    
    # Combine them with type info for template
    users_list = [{'id': u.id, 'full_name': u.full_name, 'email': u.email, 'username': u.username, 'type': 'user'} for u in all_users]
    admins_list = [{'id': a.id, 'full_name': a.full_name, 'email': a.email, 'username': a.username, 'type': 'admin'} for a in all_admins]
    
    combined_users = users_list + admins_list
    
    return render_template('admin/user_management.html', users=combined_users)


@app.route('/admin/user/<int:user_id>/edit', methods=['GET', 'POST'])
def admin_edit_user(user_id):
    """Edit user account"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    # Try to find user in User table first
    user = User.query.get(user_id)
    user_type = 'user'
    
    # If not found, try Admin table
    if not user:
        user = Admin.query.get(user_id)
        user_type = 'admin'
    
    if not user:
        flash('User not found', 'danger')
        return redirect(url_for('admin_user_management'))
    
    if request.method == 'POST':
        try:
            user.full_name = request.form.get('full_name', user.full_name)
            user.email = request.form.get('email', user.email)
            user.username = request.form.get('username', user.username)
            
            # If switching user type (only for regular users)
            if user_type == 'user':
                new_type = request.form.get('user_type', 'user')
                if new_type == 'admin' and user_type == 'user':
                    # Create new admin and delete user
                    new_admin = Admin(
                        username=user.username,
                        email=user.email,
                        full_name=user.full_name,
                        password_hash=user.password_hash
                    )
                    db.session.add(new_admin)
                    db.session.delete(user)
            
            db.session.commit()
            flash(f'✓ User {user.full_name} updated successfully', 'success')
            return redirect(url_for('admin_user_management'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error updating user: {str(e)}', 'danger')
    
    return render_template('admin/edit_user.html', user=user, user_type=user_type)


@app.route('/admin/user/<int:user_id>/delete', methods=['POST'])
def admin_delete_user(user_id):
    """Delete user account"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    # Try to find user in User table first
    user = User.query.get(user_id)
    user_model = User
    
    # If not found, try Admin table
    if not user:
        user = Admin.query.get(user_id)
        user_model = Admin
    
    if not user:
        flash('User not found', 'danger')
        return redirect(url_for('admin_user_management'))
    
    # Prevent deleting the current admin
    if user.id == session.get('user_id') and user_model == Admin:
        flash('Cannot delete your own account', 'warning')
        return redirect(url_for('admin_user_management'))
    
    try:
        user_name = user.full_name
        
        # Delete user permissions first (referential integrity) - only for regular users
        if user_model == User:
            UserAnalysisPermission.query.filter_by(user_id=user.id).delete()
        
        # Delete the user
        db.session.delete(user)
        db.session.commit()
        flash(f'✓ User {user_name} deleted successfully', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting user: {str(e)}', 'danger')
    
    return redirect(url_for('admin_user_management'))


@app.route('/admin/user/create', methods=['GET', 'POST'])
def admin_create_user():
    """Create new user account"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        try:
            email = request.form.get('email')
            password = request.form.get('password')
            full_name = request.form.get('full_name')
            username = request.form.get('username')
            user_type = request.form.get('user_type', 'user')
            
            # Check if email already exists in User table
            existing_user = User.query.filter_by(email=email).first()
            if not existing_user:
                existing_user = Admin.query.filter_by(email=email).first()
            
            if existing_user:
                flash('Email already exists', 'warning')
                return redirect(url_for('admin_create_user'))
            
            # Create new user or admin
            if user_type == 'admin':
                new_user = Admin(
                    email=email,
                    username=username,
                    full_name=full_name
                )
                new_user.set_password(password)
            else:
                new_user = User(
                    email=email,
                    username=username,
                    full_name=full_name
                )
                new_user.set_password(password)
            
            db.session.add(new_user)
            db.session.commit()
            flash(f'✓ User {full_name} created successfully', 'success')
            return redirect(url_for('admin_user_management'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error creating user: {str(e)}', 'danger')
    
    return render_template('admin/create_user.html')


# ==================== DEBUG ROUTES ====================

@app.route('/debug/user-permissions/<int:user_id>')
def debug_user_permissions(user_id):
    """Debug route to check user permissions"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return "Unauthorized", 401
    
    user = User.query.get(user_id)
    if not user:
        return f"User {user_id} not found", 404
    
    # Get all permissions for this user
    permissions = UserAnalysisPermission.query.filter_by(user_id=user_id).all()
    
    # Get all analyses shared with this user
    shared_analyses = Analysis.query.join(
        UserAnalysisPermission,
        UserAnalysisPermission.analysis_id == Analysis.id
    ).filter(
        UserAnalysisPermission.user_id == user_id
    ).all()
    
    debug_info = f"""
    <h2>Debug Info for User: {user.full_name} (ID: {user.id})</h2>
    
    <h3>Permissions in Database:</h3>
    <p>Total permissions: {len(permissions)}</p>
    <ul>
    """
    
    for perm in permissions:
        analysis = Analysis.query.get(perm.analysis_id)
        debug_info += f"<li>Permission ID: {perm.id}, Analysis ID: {perm.analysis_id}, Analysis Title: {analysis.title if analysis else 'NOT FOUND'}, Shared At: {perm.shared_at}</li>"
    
    debug_info += "</ul>"
    
    debug_info += f"""
    <h3>Analyses Via Join Query:</h3>
    <p>Total analyses found: {len(shared_analyses)}</p>
    <ul>
    """
    
    for analysis in shared_analyses:
        debug_info += f"<li>Analysis ID: {analysis.id}, Title: {analysis.title}</li>"
    
    debug_info += "</ul>"
    
    return debug_info


# ==================== API ROUTES ====================

@app.route('/api/data/latest')
def api_latest_data():
    """Get latest data"""
    if 'user_id' not in session:
        return jsonify({'error': 'Unauthorized'}), 401
    
    latest = DataFile.query.filter_by(is_active=True).order_by(DataFile.uploaded_at.desc()).first()
    
    if not latest:
        return jsonify({'error': 'No data'}), 404
    
    try:
        df = get_dataframe_from_file(latest.id)
        if df is None:
            return jsonify({'error': 'Error reading file'}), 500
        
        return jsonify({
            'data': df.to_dict('records'),
            'columns': df.columns.tolist(),
            'shape': df.shape
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/student/<int:sid>')
def api_student(sid):
    """Get student data"""
    if 'user_id' not in session or session.get('user_type') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 401
    
    latest = DataFile.query.filter_by(is_active=True).order_by(DataFile.uploaded_at.desc()).first()
    
    if not latest:
        return jsonify({'error': 'No data'}), 404
    
    try:
        df = get_dataframe_from_file(latest.id)
        if df is None:
            return jsonify({'error': 'Error reading file'}), 500
        
        if sid >= len(df):
            return jsonify({'error': 'Student not found'}), 404
        
        return jsonify({
            'id': sid,
            'data': df.iloc[sid].to_dict(),
            'total': len(df)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
        
        return jsonify({
            'id': sid,
            'data': df.iloc[sid].to_dict(),
            'total': len(df)
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ==================== UTILITY ====================

def allowed_file(filename):
    """Check if file allowed"""
    ALLOWED = {'csv', 'xlsx', 'xls'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED


# ==================== ERROR HANDLERS ====================

@app.errorhandler(404)
def not_found(e):
    return render_template('404.html'), 404

@app.errorhandler(500)
def error(e):
    return render_template('500.html'), 500


# ==================== MAIN ====================

if __name__ == '__main__':
    init_db()
    print("\n" + "="*70)
    print("COGNITIVE & EMOTIONAL PERFORMANCE SYSTEM")
    print("="*70)
    print("🌐 URL: http://localhost:5000")
    print("📝 Default Admin: admin / admin@2024")
    print("📋 Register: http://localhost:5000/register")
    print("="*70 + "\n")
    app.run(debug=True, port=5000, use_reloader=False)
